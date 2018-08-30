# -*- coding: utf-8 -*-
import os, sys
import time, datetime
import traceback
from mako.template import Template
from bson.objectid import ObjectId

sys.path.append(os.path.join(os.path.split(os.path.realpath(__file__))[0], '../util'))
import loghelper, db, email_helper, funding_helper

sys.path.append(os.path.join(os.path.split(os.path.realpath(__file__))[0], '../spider2/api/api/data'))
import data_code

# logger
loghelper.init_logger("org_track_funding_email", stream=True)
logger = loghelper.get_logger("org_track_funding_email")


def process(test=True):
    conn = db.connect_torndb()
    orgs = conn.query("select distinct organizationId from org_function_switch where "
                      "active='Y' and functionState='Y' and functionValue=68024")
    conn.close()
    for org in orgs:
        org_id = org["organizationId"]
        logger.info("orgId: %s", org_id)
        process_one(org_id, test=test)


def process_one(org_id, thedate=None, test=True):
    if thedate is None:
        today = datetime.datetime.now()
    else:
        today = thedate

    # 获取上周六
    startDate = (today - datetime.timedelta(days=time.localtime().tm_wday + 2))
    start_time = datetime.datetime(startDate.year, startDate.month, startDate.day, 21)
    # 获取这周六
    endDate = (today - datetime.timedelta(days=time.localtime().tm_wday - 5))
    end_time = datetime.datetime(endDate.year, endDate.month, endDate.day, 21)

    conn = db.connect_torndb()
    mongo = db.connect_mongo()

    df, _ = data_code.run(conn, mongo, start_time.strftime("%Y-%m-%d"), end_time.strftime("%Y-%m-%d"))
    df = df[(df.publishDateMerge >= start_time) & (df.publishDateMerge < end_time)]

    nameMap = {}
    string = u'''首次披露时间 项目名称	领域	是否国内	一句话简介	完整简介	融资详情
publishDateMerge    companyName	sector	location	brief	description	investmentDetail'''
    stringrows = string.split('\n')
    index = 0
    for column in stringrows[1].split():
        nameMap[column] = stringrows[0].split()[index]
        index += 1

    df = df.rename(columns=nameMap)

    title = "烯牛数据融资事件表（%s ~ %s）" % (start_time.strftime("%m-%d"), end_time.strftime("%m-%d"))
    fileName = "funding (%s ~ %s).xlsx" % (start_time.strftime("%m-%d"), end_time.strftime("%m-%d"))

    from openpyxl import load_workbook
    import pandas as pd

    writer = pd.ExcelWriter(fileName, engin='openpyxl')
    book = load_workbook('template/template.xlsx')

    ws = book.active
    ws['b9'] = u'数据包含了%s至%s一周的国内外融资事件。' % (start_time.strftime("%Y年%m月%d日"), end_time.strftime("%Y年%m月%d日"))
    writer.book = book
    df.to_excel(excel_writer=writer, sheet_name=u"数据明细", index=0, columns=stringrows[0].split())
    writer.save()
    writer.close()

    path = '/data/task-201606/spider2/aggregator/funding'
    path = sys.path[0]
    path = os.path.join(path, fileName)

    content = '''Hello,<br /><br />
以下是本周（%s ~ %s）披露的国内外投融资事件列表，请查收！''' % (start_time.strftime("%m-%d"), end_time.strftime("%m-%d"))

    content = '''<!DOCTYPE html><html xmlns="http://www.w3.org/1999/xhtml" xmlns:v="urn:schemas-microsoft-com:vml" xmlns:o="urn:schemas-microsoft-com:office:office"><head>  <title></title>  <!--[if !mso]><!-- -->  <meta http-equiv="X-UA-Compatible" content="IE=edge">  <!--<![endif]--><meta http-equiv="Content-Type" content="text/html; charset=UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0"><style type="text/css">  #outlook a { padding: 0; }  .ReadMsgBody { width: 100%; }  .ExternalClass { width: 100%; }  .ExternalClass * { line-height:100%; }  body { margin: 0; padding: 0; -webkit-text-size-adjust: 100%; -ms-text-size-adjust: 100%; }  table, td { border-collapse:collapse; mso-table-lspace: 0pt; mso-table-rspace: 0pt; }  img { border: 0; height: auto; line-height: 100%; outline: none; text-decoration: none; -ms-interpolation-mode: bicubic; }  p { display: block; margin: 13px 0; }</style><!--[if !mso]><!--><style type="text/css">  @media only screen and (max-width:480px) {    @-ms-viewport { width:320px; }    @viewport { width:320px; }  }</style><!--<![endif]--><!--[if mso]><xml>  <o:OfficeDocumentSettings>    <o:AllowPNG/>    <o:PixelsPerInch>96</o:PixelsPerInch>  </o:OfficeDocumentSettings></xml><![endif]--><!--[if lte mso 11]><style type="text/css">  .outlook-group-fix {    width:100% !important;  }</style><![endif]--><!--[if !mso]><!-->    <link href="https://fonts.googleapis.com/css?family=Ubuntu:300,400,500,700" rel="stylesheet" type="text/css">    <style type="text/css">        @import url(https://fonts.googleapis.com/css?family=Ubuntu:300,400,500,700);    </style>  <!--<![endif]--><style type="text/css">  @media only screen and (min-width:480px) {    .mj-column-per-100 { width:100%!important; }  }</style></head><body style="background: #FFFFFF;">    <div class="mj-container" style="background-color:#FFFFFF;"><!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0" width="600" align="center" style="width:600px;">        <tr>          <td style="line-height:0px;font-size:0px;mso-line-height-rule:exactly;">      <![endif]--><div style="margin:0px auto;max-width:600px;background:&#x4EE5;&#x4E0A;&#x4E3A;&#x672C;&#x6B21;&#x8FFD;&#x8E2A;&#x5185;&#x5BB9;&#x3002;  &#x5982;&#x6709;&#x7591;&#x95EE;&#xFF0C;&#x6B22;&#x8FCE;&#x8054;&#x7CFB;&#x6211;&#x4EEC;&#xFF1A;&#xFF09;  &#x70EF;&#x725B;&#x6570;&#x636E;&#x56E2;&#x961F;  www.xiniudata.com;"><table role="presentation" cellpadding="0" cellspacing="0" style="font-size:0px;width:100%;background:&#x4EE5;&#x4E0A;&#x4E3A;&#x672C;&#x6B21;&#x8FFD;&#x8E2A;&#x5185;&#x5BB9;&#x3002;  &#x5982;&#x6709;&#x7591;&#x95EE;&#xFF0C;&#x6B22;&#x8FCE;&#x8054;&#x7CFB;&#x6211;&#x4EEC;&#xFF1A;&#xFF09;  &#x70EF;&#x725B;&#x6570;&#x636E;&#x56E2;&#x961F;  www.xiniudata.com;" align="center" border="0"><tbody><tr><td style="text-align:center;vertical-align:top;direction:ltr;font-size:0px;padding:9px 0px 9px 0px;"><!--[if mso | IE]>      <table role="presentation" border="0" cellpadding="0" cellspacing="0">        <tr>          <td style="vertical-align:top;width:600px;">      <![endif]--><div class="mj-column-per-100 outlook-group-fix" style="vertical-align:top;display:inline-block;direction:ltr;font-size:13px;text-align:left;width:100%;"><table role="presentation" cellpadding="0" cellspacing="0" width="100%" border="0"><tbody><tr><td style="word-wrap:break-word;font-size:0px;padding:0px 20px 0px 20px;" align="center"><div style="cursor:auto;color:#000000;font-family:Ubuntu, Helvetica, Arial, sans-serif;font-size:11px;line-height:22px;text-align:center;"><p>Hi&#xA0;&#xFF0C;</p><p>&#x9644;&#x4EF6;&#x662F;&#x672C;&#x5468;&#x62AB;&#x9732;&#x7684;&#x56FD;&#x5185;&#x5916;&#x6295;&#x878D;&#x8D44;&#x4E8B;&#x4EF6;&#x5217;&#x8868;&#xFF0C;&#x8BF7;&#x67E5;&#x6536;&#xFF01;</p><p></p><p></p><p>&#x5982;&#x6709;&#x7591;&#x95EE;&#xFF0C;&#x6B22;&#x8FCE;&#x8054;&#x7CFB;&#x6211;&#x4EEC;&#xFF1A;&#xFF09;</p><p>&#x70EF;&#x725B;&#x6570;&#x636E;&#x56E2;&#x961F;</p><p><a href="http://sctrack.sc.gg/track/click/eyJtYWlsbGlzdF9pZCI6IDAsICJ0YXNrX2lkIjogIiIsICJlbWFpbF9pZCI6ICIxNTMwMTgzNjU0NDk1XzYwMTE0XzgxNjRfNDczOS5zYy0xMF85XzRfNDAtaW5ib3VuZDAkYXJ0aHVyQHhpbml1ZGF0YS5jb20iLCAic2lnbiI6ICJkNWQ5MjZhM2I3YWM3M2E2NDQwMTMwYzRlZjUzYTg1NiIsICJ1c2VyX2hlYWRlcnMiOiB7fSwgImxhYmVsIjogMCwgImxpbmsiOiAiaHR0cCUzQS8vd3d3Lnhpbml1ZGF0YS5jb20iLCAidXNlcl9pZCI6IDYwMTE0LCAiY2F0ZWdvcnlfaWQiOiAxMTI1OTh9.html" target="_blank">www.xiniudata.com</a></p><p><img src="http://www.xiniudata.com/resources/image/icon-system/verify/ios-normal.jpeg"></p><p></p></div></td></tr></tbody></table></div><!--[if mso | IE]>      </td></tr></table>      <![endif]--></td></tr></tbody></table></div><!--[if mso | IE]>      </td></tr></table>      <![endif]--></div></body></html>
    '''

    users = conn.query("select * from org_track_user "
                       "where active='Y' and orgId=%s",
                       org_id)

    for user in users:
        if user["email"] is None or user["email"].strip() == "":
            continue
        if test is True:
            if user["email"] not in ["zhlong@xiniudata.com"]:
                continue
        logger.info("%s", user["email"])
        # email_helper.send_mail("烯牛数据","烯牛数据", "noreply@xiniudata.com", user["email"], title, content)

        email_helper.send_mail_file("烯牛数据", "烯牛数据", "noreply@xiniudata.com",
                                    user["email"], title, content, path)
        # pass

    conn.close()
    mongo.close()


if __name__ == "__main__":
    if len(sys.argv) == 1:
        process(test=False)
    else:
        if sys.argv[1] == "test":
            process(test=True)
        else:
            test = False
            _test = sys.argv[3]
            if _test == "test":
                test = True
            org_id = int(sys.argv[1])
            str_thedate = sys.argv[2]
            thedate = datetime.datetime.strptime(str_thedate, '%Y-%m-%d')
            process_one(org_id, thedate, test=test)
